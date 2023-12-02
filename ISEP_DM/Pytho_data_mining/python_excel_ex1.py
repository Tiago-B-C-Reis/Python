import os
os.chdir("/Users/tiagoreis/PycharmProjects/Python/ISEP_DM/Pytho_data_mining")

import openpyxl as xl

# Create a new Excel workbook and add a worksheet
wb = xl.Workbook()
ws = wb.active
ws.title = 'first_sheet'

# Assign values to cells
ws['C1'] = 'Value_1'
ws.cell(2, 3).value = 'Value_2'  # another way to do the same thing
ws['C3'] = 'Sum = '
ws['C4'] = 'Avg = '

# Assign values to data cells
ws['D1'] = 10
ws['D2'] = 15

# Assign formulas to cells using commas as argument separators
ws['D3'] = '=SUM(D1,D2)'
ws['D4'] = '=AVERAGE(D1,D2)'

# Save the workbook to a file
wb.save('python_excel.xlsx')

print("Excel file created successfully.")
